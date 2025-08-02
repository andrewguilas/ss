import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import win32com.client as win32
import os
import app.config as config

def get_rows_from_csv(file_name):
    with open(file_name, newline="") as csvfile:
        reader = csv.DictReader(csvfile)
        return list(reader)

def write_orders_to_xlsx(file_name, header_widths, header_wrapped, rows, title, footer_rows=None):
    wb = Workbook()
    ws = wb.active

    # title
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=title)
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=config.PDF_TITLE_SIZE, bold=True)
    title_cell.alignment = Alignment(horizontal="left")

    # headers
    header_row_index = 2
    for col_index, (header, width) in enumerate(header_widths.items(), start=1):
        cell = ws.cell(row=header_row_index, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A3"

    # data
    for row_index, row_data in enumerate(rows, start=3):
        for col_index, header in enumerate(header_widths.keys(), start=1):
            value = row_data.get(header, "")
            if isinstance(value, list):
                value = ", ".join(str(item) for item in value)
            elif isinstance(value, dict):
                value = str(value)

            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(wrap_text=header in header_wrapped, horizontal="left", vertical="top")

    # footer
    if footer_rows:
        ws.append([])
        for row in footer_rows:
            ws.append(row)

    # save
    original_file_name = file_name
    for attempt in range(1, 10):
        try:
            wb.save(file_name)
        except PermissionError:
            file_name = os.path.splitext(original_file_name)[0] + f" ({attempt}).xlsx"
        else:
            break

    return file_name

def convert_xlsx_to_pdf(xlsx_file_name, pdf_file_name):
    xlsx_file_name = os.path.abspath(xlsx_file_name)
    pdf_file_name = os.path.abspath(pdf_file_name)

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False

    try:
        wb = excel.Workbooks.Open(xlsx_file_name)
        ws = wb.Worksheets(1)

        ws.PageSetup.Orientation = config.PDF_ORIENTATION
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False

        ws.PageSetup.LeftMargin = excel.InchesToPoints(config.PDF_LEFT_MARGIN)
        ws.PageSetup.RightMargin = excel.InchesToPoints(config.PDF_RIGHT_MARGIN)
        ws.PageSetup.TopMargin = excel.InchesToPoints(config.PDF_TOP_MARGIN)
        ws.PageSetup.BottomMargin = excel.InchesToPoints(config.PDF_BOTTOM_MARGIN)

        ws.PageSetup.CenterHorizontally = True
        ws.PageSetup.CenterVertically = False

        wb.ExportAsFixedFormat(0, pdf_file_name)
    finally:
        wb.Close(SaveChanges=False)
        excel.Quit()

    return pdf_file_name
