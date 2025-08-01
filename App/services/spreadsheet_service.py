import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

def get_rows_from_csv(file_name):
    with open(file_name, newline="") as csvfile:
        reader = csv.DictReader(csvfile)
        return list(reader)

def write_orders_to_xlsx(file_name, header_widths, header_wrapped, rows, title, footer_rows=None):
    wb = Workbook()
    ws = wb.active

    # title
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=title)
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=20, bold=True)
    title_cell.alignment = Alignment(horizontal="left")

    # headers
    header_row_index = 2
    for col_index, header in enumerate(header_widths.keys(), start=1):
        cell = ws.cell(row=header_row_index, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for col_index, (header, width) in enumerate(header_widths.items(), start=1):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A3"

    # data
    for row_index, row_data in enumerate(rows, start=3):
        for col_index, header in enumerate(header_widths.keys(), start=1):
            value = row_data.get(header, "")
            if isinstance(value, list):
                value = ", ".join(str(item) for item in value)
            elif isinstance(value, dict):
                value = str(value)  # or format it however you want

            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(horizontal="left")

            if header in header_wrapped:
                cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")

    # footer
    if footer_rows:
        ws.append([])
        for row in footer_rows:
            ws.append(row)

    original_file_name = file_name
    for attempt in range(1, 10):
        try:
            wb.save(file_name)
        except PermissionError:
            file_name = f"{original_file_name.replace(".xlsx", "")} ({attempt}).xlsx" 
        else:
            break

    return file_name
